# -*- coding: utf-8 -*-
"""
parse_performance.py - 微信群业绩喜报解析
用法: python parse_performance.py
默认: 自动检测 E:\weflow\wechat-deta 下最新txt，输出xlsx到同目录

输出规则:
  - 标题行: 群名 + 业绩报表 + 4月29日-5月27日
  - 文件名: 群名_业绩日报_4.29-5.27.xlsx
  - 表头深蓝底白字、日小计浅蓝底、合计中蓝底加粗
  - 列宽自适应（中文2字符）、金额千分位格式、冻结表头
  - 格式不匹配时自动诊断并提示样例
"""
import sys, re, os, glob
from collections import defaultdict

CN_NUM = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10,
          '十一':11,'十二':12,'十三':13,'十四':14,'十五':15,'十六':16,'十七':17,
          '十八':18,'十九':19,'二十':20,'二十一':21,'二十二':22,'二十三':23,
          '二十四':24,'二十五':25,'二十六':26,'二十七':27}

# ─── 日期格式化 ───
def fmt_date(d):
    parts = d.split('-')
    return f'{int(parts[1])}月{int(parts[2])}日'

# ─── 第一步：解析消息 ───
def parse_messages(file_path):
    with open(file_path, encoding='utf-8') as f:
        text = f.read()
    lines = text.split('\n')
    ts_re = r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}) '?(.+?)'?(?:\([^)]*\))?$"
    messages, cur_dt, cur_sp, cur_lines = [], None, None, []
    start = 0
    for i, line in enumerate(lines):
        if re.match(ts_re, line): start = i; break
    for line in lines[start:]:
        line = line.strip()
        m = re.match(ts_re, line)
        if m:
            if cur_dt and cur_sp:
                content = '\n'.join(cur_lines).strip()
                if content: messages.append((cur_dt, cur_sp, content))
            cur_dt = m.group(1); cur_sp = m.group(2).strip(); cur_lines = []
        elif line and cur_dt: cur_lines.append(line)
    if cur_dt and cur_sp and cur_lines:
        content = '\n'.join(cur_lines).strip()
        if content: messages.append((cur_dt, cur_sp, content))
    return messages

# ─── 第二步：提取业绩 ───
def extract_orders(messages):
    order_re = re.compile(
        r'第([一二三四五六七八九十百千]+)单[：:]\s*'
        r'(\d+(?:\.\d+)?)'
        r'\s*'
        r'(?:[（(]([^）)]*)[）)])?'
        r'([^第\d\s（(]*)'
    )
    total_re = re.compile(r'总业绩[：:]\s*(\d+(?:\.\d+)?)?')
    cash_re  = re.compile(r'实收业绩[：:]\s*(\d+(?:\.\d+)?)?')
    card_re  = re.compile(r'卡扣[：:]\s*(\d+(?:\.\d+)?)?')
    debt_re  = re.compile(r'欠款[：:]?\s*(\d+(?:\.\d+)?)?')

    records = []
    for dt, speaker, content in messages:
        if '喜报' not in content and '单' not in content: continue
        date = dt.split(' ')[0]
        totals = {'total': 0, 'cash': 0, 'card': 0, 'debt': 0}
        for pat, key in [(total_re, 'total'), (cash_re, 'cash'),
                         (card_re, 'card'), (debt_re, 'debt')]:
            m = pat.search(content)
            if m and m.group(1): totals[key] = float(m.group(1))
        orders = {}
        for m in order_re.finditer(content):
            num_str, amount = m.group(1), m.group(2)
            paren_name, bare_name = m.group(3) or '', m.group(4) or ''
            seq = CN_NUM.get(num_str, 0)
            if seq == 0: continue
            name = paren_name.strip() if paren_name else bare_name.strip()
            name = re.sub(r'[\s，,。.]$', '', name)
            if float(amount) > 0:
                orders[seq] = (seq, float(amount), name, speaker)
        for seq, (num, amt, name, sp) in sorted(orders.items()):
            records.append((date, num, amt, name, sp, totals))
    return records

# ─── 第三步：输出xlsx ───
def save_xlsx(records, path, source_name='', date_label=''):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    by_date = defaultdict(list)
    for r in records: by_date[r[0]].append(r)

    wb = Workbook(); ws = wb.active; ws.title = '业绩日报'

    hfill = PatternFill('solid', fgColor='4472C4')
    sfill = PatternFill('solid', fgColor='D9E2F3')
    tfill = PatternFill('solid', fgColor='B4C6E7')
    thin  = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    center = Alignment(horizontal='center', vertical='center')
    amt_fmt = '#,##0.00'

    # 标题行
    title_text = f'{source_name} 业绩报表  {date_label}' if date_label else f'{source_name} 业绩报表'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font = Font(bold=True, size=14, color='1F4E79')
    tc.alignment = Alignment(horizontal='center', vertical='center')

    # 表头行
    headers = ['日期', '单号', '金额', '客户/员工', '录入人']
    for j, h in enumerate(headers):
        c = ws.cell(row=2, column=j+1, value=h)
        c.font = Font(bold=True, size=11, color='FFFFFF')
        c.fill = hfill; c.alignment = center; c.border = thin

    # 数据行 + 日小计
    row_idx, grand_n, grand_t = 3, 0, 0.0
    for date in sorted(by_date):
        orders = by_date[date]
        day_n, day_t = len(orders), sum(r[2] for r in orders)
        for r in orders:
            for j, val in enumerate([r[0], r[1], r[2], r[3], r[4]]):
                c = ws.cell(row=row_idx, column=j+1, value=val)
                c.border = thin
                if j == 2: c.number_format = amt_fmt
            row_idx += 1
        for j, val in enumerate([date, f'日小计({day_n}单)', day_t, '', '']):
            c = ws.cell(row=row_idx, column=j+1, value=val)
            c.fill = sfill; c.border = thin
            if j == 2: c.number_format = amt_fmt
        row_idx += 1
        grand_n += day_n; grand_t += day_t

    # 合计行
    for j, val in enumerate(['合计', f'{grand_n}单', grand_t, '', '']):
        c = ws.cell(row=row_idx, column=j+1, value=val)
        c.fill = tfill; c.font = Font(bold=True, size=11); c.border = thin
        if j == 2: c.number_format = amt_fmt

    # 列宽自适应
    for col in range(1, len(headers)+1):
        max_w = 0
        for r in range(1, row_idx+1):
            v = str(ws.cell(row=r, column=col).value or '')
            w = sum(2 if ord(ch) > 127 else 1 for ch in v)
            max_w = max(max_w, w)
        ws.column_dimensions[get_column_letter(col)].width = min(max_w + 4, 40)

    # 冻结表头
    ws.freeze_panes = 'A3'

    wb.save(path)
    print(f'提取 {grand_n} 单，合计 {grand_t:,.2f} 元')
    print(f'已保存: {path}')

# ─── 格式诊断 ───
def diagnose_format(messages):
    print('\n' + '=' * 55)
    print('⚠ 未识别到业绩数据 — 格式诊断')
    print('=' * 55)
    amount_pat = re.compile(r'(\d{3,6}(?:\.\d{1,2})?)')
    order_pat = re.compile(r'第[一二三四五六七八九十百千]+单')
    ho, ha, samples = 0, 0, []
    for _, sp, content in messages:
        if order_pat.search(content): ho += 1
        if amount_pat.search(content): ha += 1
        if len(samples) < 5 and amount_pat.search(content):
            samples.append((sp, content[:80]))
    print(f'消息总数: {len(messages)}')
    print(f'含"第X单": {ho} 条')
    print(f'含金额数字: {ha} 条')
    if len(messages) == 0: print('\n[诊断] 导出文件为空'); return
    if ho == 0: print('\n[诊断] 未发现"第X单"关键字 - 该群可能使用不同格式')
    if ha == 0: print('\n[诊断] 未发现金额数字')
    if samples:
        print('\n--- 消息样例 ---')
        for sp, c in samples: print(f'  [{sp}] {c}')
    print('\n请发一条业绩消息样例，我来适配新格式')
    print('-' * 55)

# ─── 自动检测导出文件 ───
def auto_detect():
    export_dir = r"E:\weflow\wechat-deta"
    if os.path.isdir(export_dir):
        txts = glob.glob(os.path.join(export_dir, "*.txt"))
        if txts: return max(txts, key=os.path.getmtime)
    return None

# ─── 入口 ───
if __name__ == '__main__':
    input_file = sys.argv[1] if len(sys.argv) >= 2 else auto_detect()
    if not input_file or not os.path.exists(input_file):
        print('未找到导出文件: python parse_performance.py <input.txt>')
        sys.exit(1)

    print(f'解析: {input_file}')
    msgs = parse_messages(input_file)
    print(f'读取 {len(msgs)} 条消息')
    recs = extract_orders(msgs)
    if not recs:
        diagnose_format(msgs)
        sys.exit(0)

    # 群名
    source_name = os.path.splitext(os.path.basename(input_file))[0]
    if source_name.startswith('群聊_'): source_name = source_name[3:]

    # 日期范围
    dates = sorted(set(r[0] for r in recs))
    ds, de = fmt_date(dates[0]), fmt_date(dates[-1])
    date_label = f'{ds}-{de}'
    date_slug = date_label.replace('月', '.').replace('日', '')

    # 输出路径
    out = sys.argv[2] if len(sys.argv) >= 3 else os.path.join(
        os.path.dirname(input_file), f'{source_name}_业绩日报_{date_slug}.xlsx')

    save_xlsx(recs, out, source_name, date_label)
